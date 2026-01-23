from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from src.config import settings
from src.core.aliases import SHEET_COLUMNS_MEETINGS, SHEET_COLUMNS_TASKS, task_status_ru
from src.db.models import Item, Project
from src.db.repositories.items_repo import (
    list_by_project,
    list_inbox,
    list_overdue,
    list_today,
    list_week,
)


def _format_dt(value: datetime | None, tz: ZoneInfo) -> str:
    if value is None:
        return ""
    if value.tzinfo is None:
        value = value.replace(tzinfo=tz)
    else:
        value = value.astimezone(tz)
    return value.strftime("%Y-%m-%d %H:%M")


def _as_local(value: datetime, tz: ZoneInfo) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=tz)
    return value.astimezone(tz)


def _format_date(value: datetime | None, tz: ZoneInfo) -> str:
    if value is None:
        return ""
    return _as_local(value, tz).strftime("%Y-%m-%d")


def _format_time(value: datetime | None, tz: ZoneInfo) -> str:
    if value is None:
        return ""
    return _as_local(value, tz).strftime("%H:%M")


def _load_items(session: Session, scope: dict) -> list[Item]:
    mode = scope.get("mode", "all")
    if mode == "inbox":
        return list_inbox(session)
    if mode == "today":
        return list_today(session)
    if mode == "week":
        return list_week(session)
    if mode == "overdue":
        return list_overdue(session)
    if mode == "project":
        return list_by_project(
            session,
            project_id=scope.get("project_id"),
            project_name=scope.get("project_name"),
        )

    return list(session.scalars(select(Item).order_by(Item.created_at.desc())).all())


def export_xlsx(session: Session, scope: dict) -> str:
    tz = ZoneInfo(settings.timezone)
    items = _load_items(session, scope)

    project_map: dict[str, str] = {
        project.id: project.name for project in session.scalars(select(Project)).all()
    }
    wb = Workbook()
    ws: Worksheet = wb.active
    if ws is None:
        ws = wb.create_sheet("Задачи")
    ws.title = "Задачи"
    _build_tasks_sheet(ws, items, project_map, tz)

    ws_meetings: Worksheet = wb.create_sheet("Встречи")
    _build_meetings_sheet(ws_meetings, items, tz)

    ws_tree: Worksheet = wb.create_sheet("Дерево")
    _build_tree_sheet(ws_tree, items, project_map, tz)

    ws_sum: Worksheet = wb.create_sheet("Сводка")
    _build_summary_sheet(ws_sum, items, project_map, tz)

    if scope.get("mode") == "week":
        ws_week: Worksheet = wb.create_sheet("Неделя")
        _build_tasks_sheet(ws_week, items, project_map, tz)

        start_day = datetime.now(tz).replace(hour=0, minute=0, second=0, microsecond=0)
        end_day = start_day + timedelta(days=1)
        today_items = [
            item
            for item in items
            if _in_window(item, start_day, end_day, tz)
        ]
        ws_today: Worksheet = wb.create_sheet("Сегодня")
        _build_tasks_sheet(ws_today, today_items, project_map, tz)

    export_dir = Path("data") / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(tz).strftime("%Y%m%d_%H%M")
    filename = f"todo_export_{stamp}.xlsx"
    path = export_dir / filename
    wb.save(path)
    return str(path)


def _build_tasks_sheet(
    ws: Worksheet,
    items: list[Item],
    project_map: dict[str, str],
    tz: ZoneInfo,
) -> None:
    headers = SHEET_COLUMNS_TASKS
    ws.append(headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:I1"

    for item in items:
        if item.type != "task":
            continue
        project_name = project_map.get(item.project_id or "", "")
        due_dt = item.due_at or item.scheduled_at
        ws.append(
            [
                item.id,
                item.title,
                task_status_ru(item.status),
                _format_date(due_dt, tz),
                _format_time(due_dt, tz),
                project_name,
                "",
                _format_dt(item.created_at, tz),
                _format_dt(item.updated_at, tz),
            ]
        )

    _autosize_columns(ws)


def _build_meetings_sheet(
    ws: Worksheet,
    items: list[Item],
    tz: ZoneInfo,
) -> None:
    headers = SHEET_COLUMNS_MEETINGS
    ws.append(headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:H1"

    for item in items:
        if item.type != "meeting":
            continue
        start_dt = item.scheduled_at
        if start_dt and start_dt.tzinfo is None:
            start_dt = start_dt.replace(tzinfo=tz)
        elif start_dt:
            start_dt = start_dt.astimezone(tz)
        end_dt = None
        if start_dt and item.duration_min:
            end_dt = start_dt + timedelta(minutes=item.duration_min)
        ws.append(
            [
                item.id,
                item.title,
                _format_date(start_dt, tz) if start_dt else "",
                _format_time(start_dt, tz) if start_dt else "",
                _format_time(end_dt, tz) if end_dt else "",
                item.duration_min or "",
                task_status_ru(item.status),
                item.calendar_id or "",
            ]
        )

    _autosize_columns(ws)


def _in_window(item: Item, start: datetime, end: datetime, tz: ZoneInfo) -> bool:
    for dt_value in (item.due_at, item.scheduled_at):
        if dt_value is None:
            continue
        local = _as_local(dt_value, tz)
        if start <= local < end:
            return True
    return False


def _autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col = column_cells[0].column
        for cell in column_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 60)


def _build_tree_sheet(ws: Worksheet, items: list[Item], project_map: dict[str, str], tz: ZoneInfo) -> None:
    headers = [
        "Проект",
        "Название",
        "Статус",
        "Дата",
        "Запланировано",
        "Длительность (мин)",
        "ID",
        "Родитель",
        "Глубина",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:I1"

    by_project: dict[str | None, list[Item]] = {}
    for item in items:
        by_project.setdefault(item.project_id, []).append(item)

    for project_id, project_items in sorted(
        by_project.items(), key=lambda kv: project_map.get(kv[0] or "", "")
    ):
        project_name = project_map.get(project_id or "", "")
        children_map: dict[str | None, list[Item]] = {}
        for item in project_items:
            children_map.setdefault(item.parent_id, []).append(item)

        def _walk(parent_id: str | None, depth: int) -> None:
            for child in sorted(children_map.get(parent_id, []), key=lambda it: it.created_at):
                title = f"{'  ' * depth}• {child.title}"
                ws.append(
                    [
                        project_name,
                        title,
                        task_status_ru(child.status),
                        _format_dt(child.due_at, tz),
                        _format_dt(child.scheduled_at, tz),
                        child.duration_min if child.duration_min is not None else "",
                        child.id,
                        child.parent_id or "",
                        child.depth,
                    ]
                )
                if depth < 3:
                    _walk(child.id, depth + 1)

        _walk(None, 0)

    _autosize_columns(ws)


def _build_summary_sheet(
    ws: Worksheet,
    items: list[Item],
    project_map: dict[str, str],
    tz: ZoneInfo,
) -> None:
    headers = [
        "Проект",
        "Всего",
        "Активные",
        "Ожидают",
        "Завершены",
        "Просрочены",
        "Встречи",
        "План (мин)",
        "Факт (мин)",
        "Дельта (мин)",
        "Обновлено",
    ]
    ws.append(headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:K1"

    now = datetime.now(tz)

    summary: dict[str, dict] = {}
    for item in items:
        project_name = project_map.get(item.project_id or "", "")
        data = summary.setdefault(
            project_name,
            {
                "Total": 0,
                "Active": 0,
                "Waiting": 0,
                "Done": 0,
                "Overdue": 0,
                "Meetings": 0,
                "PlannedTotal": 0,
                "ActualTotal": 0,
                "DeltaTotal": 0,
                "UpdatedMax": None,
            },
        )
        data["Total"] += 1
        if item.status == "active":
            data["Active"] += 1
        if item.status == "waiting":
            data["Waiting"] += 1
        if item.status == "done":
            data["Done"] += 1
        if item.type == "meeting":
            data["Meetings"] += 1
        if item.planned_min is not None:
            planned = item.planned_min
        elif item.type == "meeting":
            planned = item.duration_min if item.duration_min is not None else 60
        else:
            planned = 0
        actual = item.actual_min or 0
        data["PlannedTotal"] += planned
        data["ActualTotal"] += actual
        data["DeltaTotal"] += actual - planned

        if item.due_at:
            due_at = item.due_at
            if due_at.tzinfo is None:
                due_at = due_at.replace(tzinfo=tz)
            else:
                due_at = due_at.astimezone(tz)
            if item.status == "active" and due_at < now:
                data["Overdue"] += 1

        if item.updated_at:
            updated = item.updated_at
            if updated.tzinfo is None:
                updated = updated.replace(tzinfo=tz)
            else:
                updated = updated.astimezone(tz)
            if data["UpdatedMax"] is None or updated > data["UpdatedMax"]:
                data["UpdatedMax"] = updated

    def _write_row(project_name: str, data: dict) -> None:
        ws.append(
            [
                project_name,
                data["Total"],
                data["Active"],
                data["Waiting"],
                data["Done"],
                data["Overdue"],
                data["Meetings"],
                data["PlannedTotal"],
                data["ActualTotal"],
                data["DeltaTotal"],
                _format_dt(data["UpdatedMax"], tz),
            ]
        )

    totals = {
        "Total": 0,
        "Active": 0,
        "Waiting": 0,
        "Done": 0,
        "Overdue": 0,
        "Meetings": 0,
        "PlannedTotal": 0,
        "ActualTotal": 0,
        "DeltaTotal": 0,
        "UpdatedMax": None,
    }

    for project_name in sorted(summary.keys()):
        data = summary[project_name]
        _write_row(project_name, data)
        for key in ["Total", "Active", "Waiting", "Done", "Overdue", "Meetings", "PlannedTotal", "ActualTotal", "DeltaTotal"]:
            totals[key] += data[key]
        if data["UpdatedMax"] and (
            totals["UpdatedMax"] is None or data["UpdatedMax"] > totals["UpdatedMax"]
        ):
            totals["UpdatedMax"] = data["UpdatedMax"]

    _write_row("ALL", totals)
    _autosize_columns(ws)
